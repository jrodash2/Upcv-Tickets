import json
import os
import calendar
import logging
from uuid import uuid4

from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import models
from django.db.models import ProtectedError
from django.db.models.functions import Replace
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from PIL import Image, UnidentifiedImageError
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from empleados_app.models import ConfiguracionGeneral, Empleado

from .design_engine import (
    CANVAS_HEIGHT,
    CANVAS_WIDTH,
    build_design_definition,
    build_design_editor_payload,
    build_diploma_render_context,
    ensure_design_definition,
    normalize_definition_from_elements,
)
from .forms import (
    AgregarEmpleadoCursoForm,
    AgregarParticipanteRapidoForm,
    CursoForm,
    DisenoDiplomaForm,
    EditarParticipanteCursoForm,
    FirmaForm,
    MatriculaManualParticipanteForm,
    PublicCourseRegistrationForm,
    PublicDiplomaDownloadForm,
    normalize_dpi_input,
    UbicacionDiplomaForm,
    UsuarioUbicacionDiplomaForm,
)
from .models import (
    Curso,
    CursoEmpleado,
    DisenoDiploma,
    Firma,
    UbicacionDiploma,
    UsuarioUbicacionDiploma,
)
from .notifications import send_enrollment_notification
from .utils import attach_diplomas_context, diplomas_access_required, enforce_scope_for_object, scope_queryset

logger = logging.getLogger(__name__)


# Helpers

def render_diplomas(request, template_name, context=None):
    context = context or {}
    return render(request, template_name, attach_diplomas_context(context, request))


def get_scope(request):
    return getattr(request, "diplomas_scope", {})


def get_course_or_404(request, **lookup):
    curso = get_object_or_404(Curso.objects.select_related("ubicacion", "diseno_diploma"), **lookup)
    return enforce_scope_for_object(curso, get_scope(request))


def get_design_or_404(request, **lookup):
    diseno = get_object_or_404(DisenoDiploma.objects.select_related("ubicacion"), **lookup)
    return enforce_scope_for_object(diseno, get_scope(request))


def get_signature_or_404(request, **lookup):
    firma = get_object_or_404(Firma.objects.select_related("ubicacion"), **lookup)
    return enforce_scope_for_object(firma, get_scope(request))


def get_location_or_404(request, **lookup):
    if not get_scope(request).get("is_admin"):
        raise PermissionDenied("Solo el grupo Diplomas puede administrar ubicaciones.")
    return get_object_or_404(UbicacionDiploma, **lookup)


def get_course_by_code_or_none(codigo):
    return Curso.objects.select_related("ubicacion", "diseno_diploma").filter(codigo=codigo).first()


def annotate_normalized_dpi(queryset, field_name, alias):
    normalized = Replace(models.F(field_name), models.Value(" "), models.Value(""))
    normalized = Replace(normalized, models.Value("-"), models.Value(""))
    return queryset.annotate(**{alias: normalized})


def get_employee_by_dpi_or_none(dpi):
    normalized_dpi = normalize_dpi_input(dpi)
    if not normalized_dpi:
        return None
    empleados = annotate_normalized_dpi(Empleado.objects.all(), "dpi", "normalized_dpi")
    return empleados.filter(normalized_dpi=normalized_dpi).first()


def get_participant_by_course_and_dpi_or_none(curso, dpi):
    normalized_dpi = normalize_dpi_input(dpi)
    if not curso or not normalized_dpi:
        return None
    participantes = CursoEmpleado.objects.select_related("curso", "curso__ubicacion", "curso__diseno_diploma", "empleado", "empleado__datos_basicos")
    participantes = annotate_normalized_dpi(participantes, "participante_dpi", "normalized_participante_dpi")
    participantes = annotate_normalized_dpi(participantes, "empleado__dpi", "normalized_empleado_dpi")
    return participantes.filter(curso=curso).filter(
        models.Q(normalized_participante_dpi=normalized_dpi) | models.Q(normalized_empleado_dpi=normalized_dpi)
    ).first()

def get_participant_by_course_and_dpi_or_none(curso, dpi):
    normalized_dpi = normalize_dpi_input(dpi)
    if not curso or not normalized_dpi:
        return None
    participantes = CursoEmpleado.objects.select_related("curso", "curso__ubicacion", "curso__diseno_diploma", "empleado", "empleado__datos_basicos")
    participantes = annotate_normalized_dpi(participantes, "participante_dpi", "normalized_participante_dpi")
    participantes = annotate_normalized_dpi(participantes, "empleado__dpi", "normalized_empleado_dpi")
    return participantes.filter(curso=curso).filter(
        models.Q(normalized_participante_dpi=normalized_dpi) | models.Q(normalized_empleado_dpi=normalized_dpi)
    ).first()

def build_public_course_links(request, curso):
    registration_path = f"{reverse('diplomas:public_course_registration')}?codigo_curso={curso.codigo}"
    download_path = f"{reverse('diplomas:public_diploma_download')}?codigo_curso={curso.codigo}"
    return {
        "registration_url": request.build_absolute_uri(registration_path),
        "download_url": request.build_absolute_uri(download_path),
    }


def get_public_branding_context(course=None):
    config = ConfiguracionGeneral.objects.first()
    selected_course = course
    selected_location = getattr(selected_course, "ubicacion", None) if selected_course else None
    enrollment_open, enrollment_message = get_course_enrollment_status(selected_course)
    download_allowed, download_message = get_public_course_diploma_download_status(selected_course)
    return {
        "configuracion": config,
        "selected_course": selected_course,
        "selected_location": selected_location,
        "course_enrollment_open": enrollment_open,
        "course_enrollment_message": enrollment_message,
        "course_download_allowed": download_allowed,
        "course_download_message": download_message,
    }


def get_course_enrollment_status(curso):
    if not curso:
        return False, "Debe seleccionar un curso válido."

    start_date = getattr(curso, "fecha_inicio", None)
    end_date = getattr(curso, "fecha_fin", None)
    if not start_date or not end_date:
        return False, "No se puede inscribir porque el curso no tiene fechas definidas."

    today = timezone.localdate()
    if today < start_date:
        return False, "La inscripción a este curso aún no está disponible."
    if today > end_date:
        return False, "La inscripción a este curso ha finalizado."
    return True, ""


def get_diploma_download_status(curso, mode="public"):
    if not curso:
        return False, "Debe seleccionar un curso válido."

    if mode == "internal":
        return True, ""

    if mode != "public":
        raise ValueError("mode must be 'public' or 'internal'")

    end_date = getattr(curso, "fecha_fin", None)
    if not end_date:
        return False, "No se puede descargar el diploma porque el curso no tiene fecha de finalización definida."

    today = timezone.localdate()
    if today < end_date:
        return False, "No se puede descargar el diploma porque el curso aún no ha finalizado."
    deadline = add_months_to_date(end_date, months=6)
    if today > deadline:
        return False, "El plazo disponible para descargar este diploma desde este enlace ya ha vencido."
    return True, ""


def get_public_course_diploma_download_status(curso):
    return get_diploma_download_status(curso, mode="public")


def get_course_diploma_download_status(curso):
    """
    Compatibilidad retroactiva.
    Algunas rutas/vistas antiguas aún pueden invocar este nombre histórico.
    """
    return get_diploma_download_status(curso, mode="internal")


def add_months_to_date(source_date, months):
    total_month = (source_date.month - 1) + months
    year = source_date.year + (total_month // 12)
    month = (total_month % 12) + 1
    day = min(source_date.day, calendar.monthrange(year, month)[1])
    return source_date.replace(year=year, month=month, day=day)


def trigger_course_completion_notifications(*_args, **_kwargs):
    """
    Compatibilidad defensiva:
    el dashboard ya no debe disparar correos de finalización en requests GET.
    """
    return {"sent": 0, "skipped": 0, "errors": 0}


# Dashboard

@diplomas_access_required
def diplomas_dahsboard(request):
    trigger_course_completion_notifications(request)
    scope = get_scope(request)
    cursos = scope_queryset(Curso.objects.select_related("ubicacion"), scope).order_by("-creado_en")
    firmas = scope_queryset(Firma.objects.select_related("ubicacion"), scope).order_by("-creado_en")
    disenos = scope_queryset(DisenoDiploma.objects.select_related("ubicacion"), scope).order_by("-creado_en")
    participantes = CursoEmpleado.objects.filter(curso__in=cursos)
    ubicaciones = UbicacionDiploma.objects.order_by("nombre") if scope.get("is_admin") else UbicacionDiploma.objects.filter(id=getattr(scope.get("location"), "id", None))

    context = {
        "total_cursos": cursos.count(),
        "total_firmas": firmas.count(),
        "total_disenos": disenos.count(),
        "total_participantes": participantes.count(),
        "total_ubicaciones": ubicaciones.count(),
        "cursos_recientes": cursos[:5],
        "firmas_recientes": firmas[:5],
        "disenos_recientes": disenos[:5],
    }
    return render_diplomas(request, "diplomas/dashboard.html", context)


# Ubicaciones

@diplomas_access_required
def ubicaciones_lista(request):
    if not get_scope(request).get("is_admin"):
        raise PermissionDenied
    ubicaciones = UbicacionDiploma.objects.order_by("nombre")
    return render_diplomas(request, "diplomas/ubicaciones_lista.html", {
        "ubicaciones": ubicaciones,
        "form": UbicacionDiplomaForm(),
    })


@diplomas_access_required
def crear_ubicacion(request):
    if not get_scope(request).get("is_admin"):
        raise PermissionDenied
    if request.method == "POST":
        form = UbicacionDiplomaForm(request.POST)
        if form.is_valid():
            ubicacion = form.save(commit=False)
            ubicacion.creado_por = request.user
            ubicacion.save()
            messages.success(request, "Ubicación creada correctamente.")
        else:
            messages.error(request, "No se pudo crear la ubicación.")
    return redirect("diplomas:ubicaciones_lista")


@diplomas_access_required
def editar_ubicacion(request, ubicacion_id):
    ubicacion = get_location_or_404(request, id=ubicacion_id)
    if request.method == "POST":
        form = UbicacionDiplomaForm(request.POST, instance=ubicacion)
        if form.is_valid():
            form.save()
            messages.success(request, "Ubicación actualizada correctamente.")
            return redirect("diplomas:ubicaciones_lista")
    else:
        form = UbicacionDiplomaForm(instance=ubicacion)
    return render_diplomas(request, "diplomas/editar_ubicacion.html", {"form": form, "ubicacion": ubicacion})


@diplomas_access_required
def eliminar_ubicacion(request, ubicacion_id):
    ubicacion = get_location_or_404(request, id=ubicacion_id)
    if request.method == "POST":
        try:
            ubicacion.delete()
            messages.success(request, "Ubicación eliminada correctamente.")
        except ProtectedError:
            messages.error(request, "No se puede eliminar la ubicación porque tiene registros relacionados.")
    return redirect("diplomas:ubicaciones_lista")


@diplomas_access_required
def asignaciones_ubicacion_lista(request):
    if not get_scope(request).get("is_admin"):
        raise PermissionDenied
    asignaciones = UsuarioUbicacionDiploma.objects.select_related("usuario", "ubicacion", "asignado_por").order_by("usuario__username")
    return render_diplomas(request, "diplomas/asignaciones_ubicacion_lista.html", {
        "asignaciones": asignaciones,
        "form": UsuarioUbicacionDiplomaForm(),
    })


@diplomas_access_required
def crear_asignacion_ubicacion(request):
    if not get_scope(request).get("is_admin"):
        raise PermissionDenied
    if request.method == "POST":
        form = UsuarioUbicacionDiplomaForm(request.POST)
        if form.is_valid():
            usuario = form.cleaned_data["usuario"]
            ubicacion = form.cleaned_data["ubicacion"]
            UsuarioUbicacionDiploma.objects.update_or_create(
                usuario=usuario,
                defaults={"ubicacion": ubicacion, "asignado_por": request.user},
            )
            messages.success(request, "Asignación guardada correctamente.")
        else:
            messages.error(request, "No se pudo guardar la asignación.")
    return redirect("diplomas:asignaciones_ubicacion_lista")


@diplomas_access_required
def editar_asignacion_ubicacion(request, asignacion_id):
    if not get_scope(request).get("is_admin"):
        raise PermissionDenied
    asignacion = get_object_or_404(UsuarioUbicacionDiploma, id=asignacion_id)
    if request.method == "POST":
        form = UsuarioUbicacionDiplomaForm(request.POST, instance=asignacion)
        if form.is_valid():
            form.save(assigned_by=request.user)
            messages.success(request, "Asignación actualizada correctamente.")
            return redirect("diplomas:asignaciones_ubicacion_lista")
    else:
        form = UsuarioUbicacionDiplomaForm(instance=asignacion)
    return render_diplomas(request, "diplomas/editar_asignacion_ubicacion.html", {"form": form, "asignacion": asignacion})


@diplomas_access_required
def eliminar_asignacion_ubicacion(request, asignacion_id):
    if not get_scope(request).get("is_admin"):
        raise PermissionDenied
    asignacion = get_object_or_404(UsuarioUbicacionDiploma, id=asignacion_id)
    if request.method == "POST":
        asignacion.delete()
        messages.success(request, "Asignación eliminada correctamente.")
    return redirect("diplomas:asignaciones_ubicacion_lista")


# Firmas

@diplomas_access_required
def firmas_lista(request):
    scope = get_scope(request)
    firmas = scope_queryset(Firma.objects.select_related("ubicacion"), scope).order_by("-id")
    form = FirmaForm(scope=scope)
    return render_diplomas(request, "diplomas/firmas_lista.html", {"firmas": firmas, "form": form})


@diplomas_access_required
def crear_firma(request):
    scope = get_scope(request)
    if request.method == "POST":
        form = FirmaForm(request.POST, request.FILES, scope=scope)
        if form.is_valid():
            form.save()
            messages.success(request, "Firma creada correctamente.")
        else:
            messages.error(request, "Error al crear la firma.")
    return redirect("diplomas:firmas_lista")


@diplomas_access_required
def editar_firma(request, firma_id):
    firma = get_signature_or_404(request, id=firma_id)
    scope = get_scope(request)
    if request.method == "POST":
        form = FirmaForm(request.POST, request.FILES, instance=firma, scope=scope)
        if form.is_valid():
            form.save()
            messages.success(request, "Firma actualizada correctamente.")
            return redirect("diplomas:firmas_lista")
    else:
        form = FirmaForm(instance=firma, scope=scope)
    return render_diplomas(request, "diplomas/editar_firma.html", {"form": form, "firma": firma})


@diplomas_access_required
def eliminar_firma(request, firma_id):
    firma = get_signature_or_404(request, id=firma_id)
    if request.method == "POST":
        try:
            firma.delete()
            messages.success(request, "Firma eliminada correctamente.")
        except ProtectedError:
            messages.error(request, "No se puede eliminar la firma porque está asociada a cursos.")
    return redirect("diplomas:firmas_lista")


# Diseños

@diplomas_access_required
def disenos_lista(request):
    scope = get_scope(request)
    disenos = scope_queryset(DisenoDiploma.objects.select_related("ubicacion"), scope).order_by("-id")
    form = DisenoDiplomaForm(scope=scope)
    return render_diplomas(request, "diplomas/disenos_lista.html", {"disenos": disenos, "form": form})


@diplomas_access_required
def crear_diseno(request):
    scope = get_scope(request)
    if request.method == "POST":
        form = DisenoDiplomaForm(request.POST, request.FILES, scope=scope)
        if form.is_valid():
            diseno = form.save()
            ensure_design_definition(diseno)
            messages.success(request, "Diseño de diploma creado correctamente.")
        else:
            messages.error(request, "No se pudo crear el diseño. Revise los campos.")
    return redirect("diplomas:disenos_lista")


@diplomas_access_required
def editar_diseno(request, diseno_id):
    diseno = get_design_or_404(request, id=diseno_id)
    scope = get_scope(request)
    if request.method == "POST":
        form = DisenoDiplomaForm(request.POST, request.FILES, instance=diseno, scope=scope)
        if form.is_valid():
            diseno = form.save()
            ensure_design_definition(diseno)
            messages.success(request, "Diseño actualizado correctamente.")
            return redirect("diplomas:disenos_lista")
    else:
        form = DisenoDiplomaForm(instance=diseno, scope=scope)

    return render_diplomas(request, "diplomas/editar_diseno.html", {"form": form, "diseno": diseno})


@ensure_csrf_cookie
@diplomas_access_required
def modificar_diseno_visual(request, diseno_id):
    diseno = get_design_or_404(request, id=diseno_id)
    editor_payload = build_design_editor_payload(diseno)
    definition = editor_payload["definition"]
    context = {
        "diseno": diseno,
        "elementos_json": definition,
        "preview_context_json": editor_payload["preview_context"],
        "fondo_url": definition["elements"]["fondo_diploma"]["image_url"],
        "canvas_width": CANVAS_WIDTH,
        "canvas_height": CANVAS_HEIGHT,
    }
    return render_diplomas(request, "diplomas/editor_diseno_visual.html", context)


@diplomas_access_required
def guardar_diseno_visual(request, diseno_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    diseno = get_design_or_404(request, id=diseno_id)
    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON inválido"}, status=400)

    incoming_elements = None
    if isinstance(payload, dict):
        if isinstance(payload.get("elementos"), dict):
            incoming_elements = payload["elementos"]
        elif isinstance(payload.get("elements"), dict):
            incoming_elements = payload["elements"]
        elif isinstance(payload.get("definition"), dict) and isinstance(payload["definition"].get("elements"), dict):
            incoming_elements = payload["definition"]["elements"]

    if not isinstance(incoming_elements, dict) or not incoming_elements:
        return JsonResponse({"success": False, "error": "Debe enviar un mapa válido de elementos."}, status=400)

    try:
        normalized_definition = normalize_definition_from_elements(diseno, incoming_elements)
        diseno.estilos = normalized_definition
        diseno.save(update_fields=["estilos", "actualizado_en"])
        diseno.refresh_from_db(fields=["estilos", "actualizado_en"])
    except Exception as exc:
        return JsonResponse({"success": False, "error": f"No se pudo guardar el diseño: {exc}"}, status=500)

    return JsonResponse({
        "success": True,
        "message": "Diseño guardado correctamente.",
        "elementos": diseno.estilos.get("elements", {}),
        "definition": diseno.estilos,
    })


@diplomas_access_required
def subir_imagen_diseno_visual(request, diseno_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido."}, status=405)

    diseno = get_design_or_404(request, id=diseno_id)
    uploaded_file = request.FILES.get("image")
    if not uploaded_file:
        return JsonResponse({"success": False, "error": "Debe seleccionar una imagen."}, status=400)

    allowed_extensions = {".png", ".jpg", ".jpeg", ".webp"}
    extension = os.path.splitext(uploaded_file.name or "")[1].lower()
    if extension not in allowed_extensions:
        return JsonResponse({"success": False, "error": "Formato no permitido. Use PNG, JPG, JPEG o WEBP."}, status=400)

    if not str(getattr(uploaded_file, "content_type", "")).startswith("image/"):
        return JsonResponse({"success": False, "error": "El archivo seleccionado no es una imagen válida."}, status=400)

    try:
        image_bytes = uploaded_file.read()
        Image.open(ContentFile(image_bytes)).verify()
    except (UnidentifiedImageError, OSError, ValueError):
        return JsonResponse({"success": False, "error": "No se pudo validar la imagen enviada."}, status=400)
    finally:
        uploaded_file.seek(0)

    folder_name = slugify(diseno.nombre) or f"diseno-{diseno.id}"
    filename = f"{uuid4().hex}{extension}"
    storage_path = f"diplomas/editor/{folder_name}/{filename}"
    saved_path = default_storage.save(storage_path, uploaded_file)
    file_url = default_storage.url(saved_path)

    return JsonResponse({
        "success": True,
        "message": "Imagen subida correctamente.",
        "image_url": file_url,
        "path": saved_path,
        "filename": os.path.basename(saved_path),
    })


@diplomas_access_required
def eliminar_diseno(request, diseno_id):
    diseno = get_design_or_404(request, id=diseno_id)
    if request.method == "POST":
        if diseno.cursos.exists():
            messages.error(request, "No se puede eliminar el diseño porque está asignado a uno o más cursos.")
        else:
            try:
                diseno.delete()
                messages.success(request, "Diseño eliminado correctamente.")
            except ProtectedError:
                messages.error(request, "No se puede eliminar el diseño por integridad de datos.")
    return redirect("diplomas:disenos_lista")


# Cursos

@diplomas_access_required
def cursos_lista(request):
    trigger_course_completion_notifications(request)
    scope = get_scope(request)
    cursos = scope_queryset(Curso.objects.select_related("ubicacion", "diseno_diploma"), scope).order_by("-creado_en")
    form = CursoForm(scope=scope)
    return render_diplomas(request, "diplomas/cursos_lista.html", {"cursos": cursos, "form": form})


@diplomas_access_required
def crear_curso_modal(request):
    scope = get_scope(request)
    if request.method == "POST":
        form = CursoForm(request.POST, scope=scope)
        if form.is_valid():
            form.save()
            messages.success(request, "Curso creado correctamente.")
            return redirect("diplomas:cursos_lista")
        messages.error(request, "Corrige los errores del formulario.")
    return redirect("diplomas:cursos_lista")


@diplomas_access_required
def editar_curso(request, curso_id):
    curso = get_course_or_404(request, id=curso_id)
    scope = get_scope(request)

    if request.method == "POST":
        form = CursoForm(request.POST, instance=curso, scope=scope)
        if form.is_valid():
            form.save()
            messages.success(request, "Curso actualizado correctamente.")
            return redirect("diplomas:cursos_lista")
    else:
        form = CursoForm(instance=curso, scope=scope)

    return render_diplomas(request, "diplomas/editar_curso.html", {"form": form, "curso": curso})


@diplomas_access_required
def detalle_curso(request, curso_id):
    curso = get_course_or_404(request, id=curso_id)
    trigger_course_completion_notifications(request, curso=curso)
    participantes = CursoEmpleado.objects.filter(curso=curso).select_related("empleado", "empleado__datos_basicos")
    total_participantes = participantes.count()
    public_links = build_public_course_links(request, curso)
    can_enroll, enrollment_message = get_course_enrollment_status(curso)
    can_download, _download_message = get_public_course_diploma_download_status(curso)
    participants_for_edit = [
        {
            "participant": participante,
            "form": EditarParticipanteCursoForm(instance=participante, prefix=f"edit_{participante.id}"),
        }
        for participante in participantes
    ]

    return render_diplomas(request, "diplomas/detalle_curso.html", {
        "curso": curso,
        "participantes": participantes,
        "participants_for_edit": participants_for_edit,
        "total_participantes": total_participantes,
        "public_registration_url": public_links["registration_url"],
        "public_diploma_download_url": public_links["download_url"],
        "matricula_rapida_form": AgregarParticipanteRapidoForm(
            scope=get_scope(request),
            course=curso,
            initial={"curso": curso},
        ),
        "matricula_manual_form": MatriculaManualParticipanteForm(
            scope=get_scope(request),
            course=curso,
            initial={"curso": curso},
        ),
        "can_enroll": can_enroll,
        "enrollment_message": enrollment_message,
        "can_download": can_download,
    })


@diplomas_access_required
def exportar_participantes_excel(request, curso_id):
    curso = get_course_or_404(request, id=curso_id)
    participantes = (
        CursoEmpleado.objects.filter(curso=curso)
        .select_related("empleado", "empleado__datos_basicos")
        .order_by("fecha_asignacion", "id")
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Participantes"

    headers = [
        "No.",
        "DPI",
        "Nombre",
        "Correo",
        "Teléfono",
        "Observaciones",
        "Fecha de asignación",
        "Tipo de registro",
        "ID Empleado",
        "Foto",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for index, participante in enumerate(participantes, start=1):
        sheet.append(
            [
                index,
                participante.dpi_participante,
                participante.nombre_participante,
                participante.correo_participante,
                participante.telefono_participante,
                participante.observaciones_participante,
                timezone.localtime(participante.fecha_asignacion).strftime("%Y-%m-%d %H:%M"),
                "Empleado" if participante.empleado_id else "Manual",
                participante.empleado_id or "",
                participante.foto_participante_url,
            ]
        )

    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column].width = min(max_length + 2, 60)

    safe_name = slugify(curso.nombre) or "curso"
    filename = f"participantes_{curso.codigo}_{safe_name}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@diplomas_access_required
def editar_participante_detalle(request, curso_id, participante_id):
    curso = get_course_or_404(request, id=curso_id)
    participante = get_object_or_404(CursoEmpleado, id=participante_id, curso=curso)

    if request.method != "POST":
        return redirect("diplomas:detalle_curso", curso_id=curso.id)

    form = EditarParticipanteCursoForm(
        request.POST,
        request.FILES,
        instance=participante,
        prefix=f"edit_{participante.id}",
    )

    if not form.is_valid():
        for _, errors in form.errors.items():
            for error in errors:
                messages.error(request, error)
        return redirect("diplomas:detalle_curso", curso_id=curso.id)

    participante_editado = form.save(commit=False)
    participante_editado.participante_correo = form.cleaned_data.get("participante_correo", "") or ""
    participante_editado.participante_telefono = form.cleaned_data.get("participante_telefono", "") or ""
    participante_editado.observaciones = form.cleaned_data.get("observaciones", "") or ""
    participante_editado.save()

    messages.success(request, "Participante actualizado correctamente.")
    return redirect("diplomas:detalle_curso", curso_id=curso.id)


@diplomas_access_required
def eliminar_participante(request, curso_id, participante_id):
    curso = get_course_or_404(request, id=curso_id)
    asignacion = get_object_or_404(CursoEmpleado, id=participante_id, curso=curso)
    asignacion.delete()
    messages.success(request, "Participante eliminado del curso.")
    return redirect("diplomas:detalle_curso", curso_id=curso.id)


@diplomas_access_required
def agregar_empleado_a_curso(request):
    scope = get_scope(request)
    if request.method == "POST":
        form = AgregarEmpleadoCursoForm(request.POST, scope=scope)
        if form.is_valid():
            curso = form.cleaned_data["curso"]
            enforce_scope_for_object(curso, scope)
            can_enroll, enrollment_message = get_course_enrollment_status(curso)
            if not can_enroll:
                messages.error(request, enrollment_message)
                return redirect("diplomas:agregar_empleado_curso")
            empleado = form.cleaned_data["empleado"]

            if CursoEmpleado.objects.filter(curso=curso, empleado=empleado).exists():
                messages.warning(request, "Este empleado ya está asignado a este curso.")
                return redirect("diplomas:agregar_empleado_curso")

            participante = CursoEmpleado.objects.create(curso=curso, empleado=empleado)
            send_enrollment_notification(participante, request=request)
            messages.success(request, "Empleado agregado correctamente al curso.")
            return redirect("diplomas:agregar_empleado_curso")
    else:
        form = AgregarEmpleadoCursoForm(scope=scope)

    return render_diplomas(request, "diplomas/agregar_empleado_curso.html", {"form": form})


@diplomas_access_required
def agregar_empleado_detalle(request, curso_id):
    curso = get_course_or_404(request, id=curso_id)
    can_enroll, enrollment_message = get_course_enrollment_status(curso)
    if not can_enroll:
        messages.error(request, enrollment_message)
        return redirect("diplomas:detalle_curso", curso_id=curso.id)
    mode = request.POST.get("enrollment_mode", "manual")

    if mode == "quick":
        form = AgregarParticipanteRapidoForm(request.POST, scope=get_scope(request), course=curso)
        if not form.is_valid():
            for _, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)
            return redirect("diplomas:detalle_curso", curso_id=curso.id)

        empleado = form.cleaned_data["empleado"]
        if CursoEmpleado.objects.filter(curso=curso, empleado=empleado).exists():
            messages.warning(request, "El participante ya está inscrito en este curso.")
            return redirect("diplomas:detalle_curso", curso_id=curso.id)

        participante = CursoEmpleado.objects.create(
            curso=curso,
            empleado=empleado,
            participante_dpi=empleado.dpi,
            participante_nombre=f"{empleado.nombres} {empleado.apellidos}".strip(),
            fecha_asignacion=timezone.now(),
        )
        send_enrollment_notification(participante, request=request)
        messages.success(request, "Participante existente agregado correctamente al curso.")
        return redirect("diplomas:detalle_curso", curso_id=curso.id)

    form = MatriculaManualParticipanteForm(request.POST, request.FILES, scope=get_scope(request), course=curso)
    if not form.is_valid():
        for _, errors in form.errors.items():
            for error in errors:
                messages.error(request, error)
        return redirect("diplomas:detalle_curso", curso_id=curso.id)

    dpi = form.cleaned_data["participante_dpi"]
    nombre = form.cleaned_data["participante_nombre"]
    empleado = Empleado.objects.filter(dpi=dpi).first()

    if empleado and CursoEmpleado.objects.filter(curso=curso, empleado=empleado).exists():
        messages.warning(request, "El participante ya está inscrito en este curso.")
        return redirect("diplomas:detalle_curso", curso_id=curso.id)

    participante = form.save(commit=False)
    participante.curso = curso
    participante.fecha_asignacion = timezone.now()
    participante.participante_dpi = dpi
    participante.participante_nombre = nombre
    participante.participante_correo = form.cleaned_data.get("participante_correo", "") or ""
    participante.participante_telefono = form.cleaned_data.get("participante_telefono", "") or ""
    participante.observaciones = form.cleaned_data.get("observaciones", "") or ""
    participante.empleado = empleado

    participante.save()
    send_enrollment_notification(participante, request=request)
    messages.success(
        request,
        "Participante agregado correctamente al curso."
        if empleado
        else "Participante manual agregado correctamente al curso.",
    )
    return redirect("diplomas:detalle_curso", curso_id=curso.id)


@diplomas_access_required
def buscar_empleado_por_dpi(request):
    dpi = normalize_dpi_input(request.GET.get("dpi"))
    if not dpi:
        return JsonResponse({"error": "No se envió DPI"}, status=400)

    empleado = get_employee_by_dpi_or_none(dpi)
    if not empleado:
        return JsonResponse({"existe": False})
    return JsonResponse({
        "existe": True,
        "empleado_id": empleado.id,
        "nombres": empleado.nombres,
        "apellidos": empleado.apellidos,
        "nombre_completo": f"{empleado.nombres} {empleado.apellidos}",
        "dpi_normalizado": normalize_dpi_input(getattr(empleado, "dpi", "")),
        "foto_url": empleado.imagen.url if empleado.imagen else "",
    })


def public_buscar_curso_por_codigo(request):
    codigo = "".join(str(request.GET.get("codigo_curso") or request.GET.get("codigo") or "").split())
    if not codigo:
        return JsonResponse({"existe": False, "error": "Debe indicar un código de curso."}, status=400)

    curso = get_course_by_code_or_none(codigo)
    if not curso:
        return JsonResponse({"existe": False})

    return JsonResponse({
        "existe": True,
        "curso_id": curso.id,
        "codigo": curso.codigo,
        "nombre": curso.nombre,
        "ubicacion": getattr(curso.ubicacion, "nombre", ""),
        "ubicacion_abreviatura": getattr(curso.ubicacion, "abreviatura", ""),
    })


def public_buscar_participante_por_dpi(request):
    codigo = "".join(str(request.GET.get("codigo_curso") or "").split())
    dpi = normalize_dpi_input(request.GET.get("dpi"))
    if not codigo or not dpi:
        return JsonResponse({"existe": False, "error": "Debe indicar código de curso y DPI."}, status=400)

    curso = get_course_by_code_or_none(codigo)
    if not curso:
        return JsonResponse({"existe": False, "error": "No existe un curso con ese código."}, status=404)

    participante = get_participant_by_course_and_dpi_or_none(curso, dpi)
    if participante:
        return JsonResponse({
            "existe": True,
            "inscrito_en_curso": True,
            "nombre_completo": participante.nombre_participante,
            "dpi": participante.dpi_participante,
            "correo": participante.correo_participante,
            "telefono": participante.telefono_participante,
        })

    empleado = get_employee_by_dpi_or_none(dpi)
    if not empleado:
        return JsonResponse({"existe": False, "inscrito_en_curso": False})

    return JsonResponse({
        "existe": True,
        "inscrito_en_curso": False,
        "nombre_completo": f"{empleado.nombres} {empleado.apellidos}".strip(),
        "dpi": empleado.dpi,
        "foto_url": empleado.imagen.url if empleado.imagen else "",
    })


def public_course_registration(request):
    initial_course_code = "".join(str(request.GET.get("codigo_curso") or request.GET.get("codigo") or "").split())
    initial_course = get_course_by_code_or_none(initial_course_code) if initial_course_code else None
    active_course = initial_course
    initial_data = {}
    if initial_course:
        initial_data = {
            "codigo_curso": initial_course.codigo,
            "nombre_curso": initial_course.nombre,
        }

    form = PublicCourseRegistrationForm(request.POST or None, request.FILES or None, initial=initial_data or None)
    registration_result = None

    if request.method == "POST" and form.is_valid():
        codigo = form.cleaned_data["codigo_curso"]
        dpi = form.cleaned_data["dpi"]
        curso = get_course_by_code_or_none(codigo)
        active_course = curso

        if not curso:
            form.add_error("codigo_curso", "No existe un curso con ese código.")
        else:
            can_enroll, enrollment_message = get_course_enrollment_status(curso)
            if not can_enroll:
                form.add_error(None, enrollment_message)
                context = {
                    "form": form,
                    "registration_result": registration_result,
                }
                context.update(get_public_branding_context(active_course or getattr(registration_result, "curso", None)))
                return render(request, "diplomas/public_course_registration.html", context)

        if curso and get_participant_by_course_and_dpi_or_none(curso, dpi):
            form.add_error("dpi", "Este participante ya está inscrito en el curso.")
        elif curso:
            empleado = get_employee_by_dpi_or_none(dpi)
            nombre = form.cleaned_data.get("participante_nombre") or ""
            if empleado:
                nombre = f"{empleado.nombres} {empleado.apellidos}".strip()
            if not nombre.strip():
                form.add_error("participante_nombre", "Debe ingresar el nombre del participante si el DPI no existe.")
            else:
                participante = CursoEmpleado(
                    curso=curso,
                    empleado=empleado,
                    participante_dpi=dpi,
                    participante_nombre=nombre.strip(),
                    participante_correo=form.cleaned_data.get("participante_correo", "") or "",
                    participante_telefono=form.cleaned_data.get("participante_telefono", "") or "",
                    observaciones=form.cleaned_data.get("observaciones", "") or "",
                    fecha_asignacion=timezone.now(),
                )
                foto = form.cleaned_data.get("participante_foto")
                if foto:
                    participante.participante_foto = foto
                participante.save()
                send_enrollment_notification(participante, request=request)
                registration_result = participante
                form = PublicCourseRegistrationForm(
                    initial={
                        "codigo_curso": curso.codigo,
                        "nombre_curso": curso.nombre,
                        "dpi": participante.dpi_participante,
                        "nombre_existente": participante.nombre_participante,
                    }
                )

    context = {
        "form": form,
        "registration_result": registration_result,
    }
    context.update(get_public_branding_context(active_course or getattr(registration_result, "curso", None)))
    return render(request, "diplomas/public_course_registration.html", context)


def public_diploma_download(request):
    initial_course_code = "".join(str(request.GET.get("codigo_curso") or request.GET.get("codigo") or "").split())
    initial_dpi = normalize_dpi_input(request.GET.get("dpi"))
    initial_course = get_course_by_code_or_none(initial_course_code) if initial_course_code else None
    active_course = initial_course
    initial_data = {}
    if initial_course:
        initial_data = {
            "codigo_curso": initial_course.codigo,
            "nombre_curso": initial_course.nombre,
        }
    if initial_dpi:
        initial_data["dpi"] = initial_dpi

    form = PublicDiplomaDownloadForm(request.POST or None, initial=initial_data or None)
    participant = None

    if request.method == "POST" and form.is_valid():
        codigo = form.cleaned_data["codigo_curso"]
        dpi = form.cleaned_data["dpi"]
        curso = get_course_by_code_or_none(codigo)
        active_course = curso

        if not curso:
            form.add_error("codigo_curso", "No existe un curso con ese código.")
        else:
            can_download, download_message = get_public_course_diploma_download_status(curso)
            if not can_download:
                form.add_error(None, download_message)
                context = {
                    "form": form,
                    "participant": participant,
                }
                context.update(get_public_branding_context(active_course or getattr(participant, "curso", None)))
                return render(request, "diplomas/public_diploma_download.html", context)

            participant = get_participant_by_course_and_dpi_or_none(curso, dpi)
            if not participant:
                employee = get_employee_by_dpi_or_none(dpi)
                if employee:
                    form.add_error("dpi", "El participante existe, pero no está inscrito en ese curso.")
                else:
                    form.add_error("dpi", "No existe un participante con el DPI indicado.")
            else:
                context = build_diploma_render_context(participant)
                context["allow_download"] = True
                context["download_block_message"] = ""
                return render(request, "diplomas/ver_diploma.html", context)

    context = {
        "form": form,
        "participant": participant,
    }
    context.update(get_public_branding_context(active_course or getattr(participant, "curso", None)))
    return render(request, "diplomas/public_diploma_download.html", context)

    form = PublicDiplomaDownloadForm(request.POST or None, initial=initial_data or None)
    participant = None

@diplomas_access_required
def ver_diploma(request, curso_id, participante_id):
    curso_empleado = get_object_or_404(CursoEmpleado.objects.select_related("curso", "curso__ubicacion", "empleado"), id=participante_id, curso_id=curso_id)
    enforce_scope_for_object(curso_empleado.curso, get_scope(request))
    can_download, download_message = get_course_diploma_download_status(curso_empleado.curso)
    if not can_download:
        messages.error(request, download_message)
        return redirect("diplomas:detalle_curso", curso_id=curso_empleado.curso_id)
    context = build_diploma_render_context(curso_empleado)
    context["allow_download"] = True
    context["download_block_message"] = ""
    return render_diplomas(request, "diplomas/ver_diploma.html", context)


@csrf_exempt
@diplomas_access_required
def guardar_posiciones(request, curso_id):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    curso = get_course_or_404(request, id=curso_id)

    try:
        data = json.loads(request.body)
    except Exception as e:
        return JsonResponse({"error": f"JSON inválido: {str(e)}"}, status=400)

    posiciones_limpias = {}
    for key, values in data.items():
        posiciones_limpias[key] = {
            "left": int(values.get("left", 0)),
            "top": int(values.get("top", 0)),
            "width": int(values.get("width", 0)),
            "height": int(values.get("height", 0)),
            "scale": float(values.get("scale", 1)),
        }

    if curso.diseno_diploma:
        if curso.diseno_diploma.ubicacion_id and curso.ubicacion_id != curso.diseno_diploma.ubicacion_id:
            return JsonResponse({"error": "El diseño del curso no coincide con su ubicación."}, status=400)
        current_definition = build_design_definition(curso.diseno_diploma, None)
        patched_elements = current_definition["elements"]
        for key, values in posiciones_limpias.items():
            if key not in patched_elements:
                continue
            patched_elements[key]["x"] = values["left"]
            patched_elements[key]["y"] = values["top"]
            patched_elements[key]["width"] = values["width"]
            patched_elements[key]["height"] = values["height"]

        curso.diseno_diploma.estilos = normalize_definition_from_elements(curso.diseno_diploma, patched_elements)
        curso.diseno_diploma.save(update_fields=["estilos", "actualizado_en"])
    else:
        curso.posiciones = posiciones_limpias
        curso.save(update_fields=["posiciones"])

    return JsonResponse({"success": True})
