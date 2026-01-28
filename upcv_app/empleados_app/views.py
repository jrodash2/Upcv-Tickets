from django.shortcuts import render, redirect, get_object_or_404
from .forms import ContratoForm, EmpleadoForm, EmpleadoeditForm, PuestoForm, SedeForm
from django.contrib.auth.models import User
from django.contrib import messages
from .models import Contrato, Empleado, Puesto
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, logout, login as auth_login  
from django.shortcuts import render, redirect
from django.contrib.auth.forms import AuthenticationForm
from .models import ConfiguracionGeneral
from .forms import ConfiguracionGeneralForm
from django.urls import reverse
from django.http import JsonResponse



from PIL import Image
from io import BytesIO
from django.template.loader import render_to_string

from django.db.models.functions import ExtractYear
from django.db.models import Count

import qrcode

from django.http import HttpResponse
from django.core.files.base import ContentFile


from django.shortcuts import render, get_object_or_404
from .models import Empleado, DatosBasicosEmpleado, FormacionAcademicaEmpleado
from .forms import DatosBasicosEmpleadoForm, FormacionAcademicaEmpleadoForm
from django.http import JsonResponse
from django.template.loader import render_to_string


def perfil_empleado(request, empleado_id):

    empleado = get_object_or_404(Empleado, id=empleado_id)

    # Formulario para modal datos básicos
    form_datos_basicos = DatosBasicosEmpleadoForm(
        instance=getattr(empleado, 'datos_basicos', None)
    )

    # Formulario para agregar formación
    form_formacion = FormacionAcademicaEmpleadoForm()

    # Todas las formaciones
    formaciones = empleado.formaciones.all()

    # Añadir formulario de edición inline
    for f in formaciones:
        f.form_editar = FormacionAcademicaEmpleadoForm(instance=f)

    return render(request, "empleados/perfil_empleado.html", {
        "empleado": empleado,
        "form_datos_basicos": form_datos_basicos,
        "form_formacion": form_formacion,
        "formaciones": formaciones,
    })



def guardar_datos_basicos(request, empleado_id):
    empleado = get_object_or_404(Empleado, id=empleado_id)
    instancia = getattr(empleado, 'datos_basicos', None)

    form = DatosBasicosEmpleadoForm(request.POST, instance=instancia)

    if form.is_valid():
        datos = form.save(commit=False)
        datos.empleado = empleado
        datos.save()

        # HTML generado inline
        html = f"""
        <div class='row'>
            <div class='col-md-6'><strong>Fecha nacimiento:</strong> {datos.fecha_nacimiento}</div>
            <div class='col-md-6'><strong>Sexo:</strong> {datos.get_sexo_display() if datos.sexo else ''}</div>

            <div class='col-md-6'><strong>Estado civil:</strong> {datos.get_estado_civil_display() if datos.estado_civil else ''}</div>
            <div class='col-md-6'><strong>Nacionalidad:</strong> {datos.nacionalidad or ''}</div>

            <div class='col-md-6'><strong>Grupo étnico:</strong> {datos.grupo_etnico or ''}</div>
            <div class='col-md-6'><strong>Idiomas:</strong> {datos.idiomas or ''}</div>

            <div class='col-md-12'><strong>Dirección:</strong> {datos.direccion_residencia or ''}</div>

            <div class='col-md-6'><strong>Teléfono personal:</strong> {datos.telefono_personal or ''}</div>
            <div class='col-md-6'><strong>Teléfono emergencia:</strong> {datos.telefono_emergencia or ''}</div>

            <div class='col-md-12'><strong>Contacto emergencia:</strong> {datos.persona_contacto_emergencia or ''}</div>
            <div class='col-md-12'><strong>Correo institucional:</strong> {datos.correo_institucional or ''}</div>
        </div>
        """

        return JsonResponse({
            "status": "reload",
            "update_section": "datos_basicos",
            "html": html
        })

    # ⚠️ IMPORTANTE: devolver errores si el form NO es válido
    return JsonResponse({
        "status": "error",
        "errors": form.errors,
    }, status=400)

def guardar_formacion(request, empleado_id):
    empleado = get_object_or_404(Empleado, id=empleado_id)
    form = FormacionAcademicaEmpleadoForm(request.POST)

    if form.is_valid():
        f = form.save(commit=False)
        f.empleado = empleado
        f.save()

        # 🔵 Aquí pedimos recargar página
        return JsonResponse({
            "status": "reload"
        })

    return JsonResponse({'status': 'error', 'msg': 'Formulario no válido'}, status=400)


def actualizar_formacion(request, formacion_id):
    formacion = get_object_or_404(FormacionAcademicaEmpleado, id=formacion_id)
    form = FormacionAcademicaEmpleadoForm(request.POST, instance=formacion)

    if form.is_valid():
        form.save()
        return refrescar_tabla_formaciones(formacion.empleado)

    return JsonResponse({'status': 'error', 'msg': 'Formulario no válido'}, status=400)


def eliminar_formacion(request, formacion_id):
    formacion = get_object_or_404(FormacionAcademicaEmpleado, id=formacion_id)
    empleado = formacion.empleado
    formacion.delete()
    return refrescar_tabla_formaciones(empleado)



def refrescar_tabla_formaciones(empleado):
    filas = ""
    for f in empleado.formaciones.all():
        filas += f"""
        <tr>
            <td>{f.get_nivel_display()}</td>
            <td>{f.titulo_obtenido or ""}</td>
            <td>{f.centro_estudio}</td>
            <td>{f.fecha}</td>
            <td class="text-end">
                <button class="btn btn-warning btn-sm" data-bs-toggle="modal"
                        data-bs-target="#modalFormacionEditar{f.id}">
                    Editar
                </button>
                <button class="btn btn-danger btn-sm btn-eliminar-formacion"
                        data-url="/empleados/formacion/{f.id}/eliminar/">
                    Eliminar
                </button>
            </td>
        </tr>
        """

    html = f"""
    <table class="table table-striped">
      <thead>
        <tr>
          <th>Nivel</th>
          <th>Título</th>
          <th>Centro</th>
          <th>Fecha</th>
          <th>Acciones</th>
        </tr>
      </thead>
      <tbody>
        {filas}
      </tbody>
    </table>
    """

    return JsonResponse({
        'status': 'ok',
        'update_section': 'formaciones',
        'html': html
    })



def crear_contrato(request, empleado_id):
    empleado = get_object_or_404(Empleado, id=empleado_id)

    if request.method == 'POST':
        form = ContratoForm(request.POST, request.FILES)
        if form.is_valid():
            contrato = form.save(commit=False)
            contrato.empleado = empleado
            contrato.save()
            return redirect('empleados:contratos', empleado_id=empleado.id)
    else:
        form = ContratoForm()

    return render(request, 'empleados/crear_contrato.html', {
        'form': form,
        'empleado': empleado,
        'sede_form': SedeForm(),
        'puesto_form': PuestoForm()
    })



def crear_sede(request):
    if request.method == 'POST':
        form = SedeForm(request.POST)
        if form.is_valid():
            form.save()
    return redirect(request.META.get('HTTP_REFERER', '/'))

def crear_puesto(request):
    if request.method == 'POST':
        form = PuestoForm(request.POST)
        if form.is_valid():
            form.save()
    return redirect(request.META.get('HTTP_REFERER', '/'))

def contratos(request, empleado_id):
    empleado = get_object_or_404(Empleado, id=empleado_id)
    contratos = empleado.contratos.all().order_by('-fecha_inicio')  # últimos contratos primero

    return render(request, 'empleados/contratos.html', {
        'empleado': empleado,
        'contratos': contratos
    })

def obtener_puestos_por_sede(request):
    sede_id = request.GET.get('sede_id')
    puestos = Puesto.objects.filter(sede_id=sede_id).values('id', 'nombre')
    return JsonResponse(list(puestos), safe=False)

def configuracion_general(request):
    configuracion, created = ConfiguracionGeneral.objects.get_or_create(id=1)  # Solo una configuración general
    if request.method == 'POST':
        form = ConfiguracionGeneralForm(request.POST, request.FILES, instance=configuracion)
        if form.is_valid():
            form.save()
            return redirect('empleados:configuracion_general')  # Redirige al formulario para ver los cambios
    else:
        form = ConfiguracionGeneralForm(instance=configuracion)
    
    return render(request, 'empleados/configuracion_general.html', {'form': form, 'configuracion': configuracion})

def home(request):
    return render(request, 'empleados/login.html')

from django.utils.timezone import now



@login_required
def dahsboard(request):
    # Empleados activos/inactivos
    total_activos = Empleado.objects.filter(activo=True).count()
    total_inactivos = Empleado.objects.filter(activo=False).count()

    # Contratos vigentes y no vigentes
    fecha_actual = now()
    contratos_vigentes = Contrato.objects.filter(
        activo=True,
        fecha_inicio__lte=fecha_actual,
        fecha_vencimiento__gte=fecha_actual
    ).count()

    contratos_no_vigentes = Contrato.objects.exclude(
        activo=True,
        fecha_inicio__lte=fecha_actual,
        fecha_vencimiento__gte=fecha_actual
    ).count()

    # Contratos por año
    contratos_por_anio = (
        Contrato.objects.annotate(anio=ExtractYear('fecha_inicio'))
        .values('anio')
        .annotate(total=Count('id'))
        .order_by('anio')
    )

    # Empleados por sede (usando contrato.sede)
    empleados_por_sede = (
    Contrato.objects
    .filter(activo=True, sede__isnull=False)
    .values('sede__nombre')
    .annotate(total=Count('empleado', distinct=True))
    .order_by('sede__nombre')
)


    print("Empleados por sede:", list(empleados_por_sede))  # <-- para debug


    # Últimos contratos
    ultimos_contratos = (
        Contrato.objects.select_related('empleado')
        .order_by('-created_at')[:4]
    )

    context = {
        'datos_empleados': {
            'activos': total_activos,
            'inactivos': total_inactivos,
            'contratos_vigentes': contratos_vigentes,
            'contratos_no_vigentes': contratos_no_vigentes,
        },
        'contratos_por_anio': list(contratos_por_anio),
        'empleados_por_sede': list(empleados_por_sede),
        'ultimos_contratos': ultimos_contratos,
    }

    return render(request, 'empleados/dahsboard.html', context)



def signout(request):
    logout(request)
    return redirect('empleados:signin')

def signin(request):  
    if request.method == 'GET':
        return render(request, 'empleados/login.html', {
            'form': AuthenticationForm
        })
    else:
       
        user = authenticate(
            request, username=request.POST['username'], password=request.POST['password']
        )
        if user is None:
            return render(request, 'empleados/login.html', {
                'form': AuthenticationForm,
                'error': 'Usuario o Password es Incorrecto'
            })
        else:
            
            auth_login(request, user)  
                      
            data = user.groups.all()
            for g in data:
                print(g.name)
                if g.name == 'Admin_gafetes':
                    return redirect('empleados:dahsboard')
                elif g.name == 'Admin_tickets':
                    return redirect('tickets:dashboard')
                elif g.name == 'tecnico':
                    return redirect('tickets:tickets_dahsboard')
                elif g.name == 'Diplomas':
                    return redirect('diplomas:diplomas_dahsboard')
                if g.name in ['Administrador', 'PRESUPUESTO']:
                     return redirect('scompras:dahsboard')
                elif g.name == 'scompras':
                    return redirect('scompras:dashboard_usuario')
                else:
                    return redirect('dahsboard')

from django.http import JsonResponse
from .models import Empleado   # Asegúrate de que el modelo está aquí
def buscar_empleado_dpi(request):
    dpi = request.GET.get("dpi", "").strip()

    if not dpi:
        return JsonResponse({"error": "Debe proporcionar un DPI."}, status=400)

    try:
        empleado = Empleado.objects.get(dpi=dpi)
    except Empleado.DoesNotExist:
        return JsonResponse({"error": "Empleado no encontrado."}, status=404)

    # Generar automáticamente username recomendado
    username = (empleado.nombres.split()[0][0] + empleado.apellidos.replace(" ", "")).lower()

    return JsonResponse({
        "nombres": empleado.nombres,
        "apellidos": empleado.apellidos,
        "foto": empleado.imagen.url if empleado.imagen else None,
        "username": username
    })



@login_required
def crear_empleado(request):
    if request.method == 'POST':
        form = EmpleadoForm(request.POST, request.FILES)
        if form.is_valid():
            empleado = form.save(commit=False)  # Crea el objeto pero no lo guarda aún
            empleado.user = request.user  # Asignamos el usuario logueado
            empleado.save()  # Ahora sí lo guardamos en la base de datos
            return redirect('empleados:empleado_lista')  # Redirige a la lista de empleados
    else:
        form = EmpleadoForm()

    return render(request, 'empleados/crear_empleado.html', {'form': form})

  
@login_required
def editar_empleado(request, e_id):
    empleado = get_object_or_404(Empleado, pk=e_id)

    if request.method == 'POST':
        form = EmpleadoeditForm(request.POST, request.FILES, instance=empleado)
        if form.is_valid():
            form.save()
            return redirect('empleados:empleado_lista')
    else:
        form = EmpleadoeditForm(instance=empleado)

    return render(request, 'empleados/editar_empleado.html', {
        'form': form,
        'empleado': empleado   # 🔵 AHORA SÍ
    })





@login_required
def lista_empleados(request):
    empleados = Empleado.objects.exclude(contratos__activo=True).distinct()

    # Obtener el último contrato por empleado (el más reciente por fecha_inicio)
    for empleado in empleados:
        empleado.ultimo_contrato = empleado.contratos.order_by('-fecha_inicio').first()

    return render(request, 'empleados/lista_empleados.html', {
        'empleados': empleados
    })


@login_required 
def credencial_empleados(request):
    empleados = Empleado.objects.filter(contratos__activo=True).distinct()  
    return render(request, 'empleados/credencial_empleados.html', {'empleados': empleados})

import openpyxl
from django.http import HttpResponse
from empleados_app.models import Empleado

def exportar_empleados_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados y Contratos"

    headers = [
        "Nombres", "Apellidos", "DPI", "Cargo", "Detalle del Cargo", "Renglón",
        "Sede", "Puesto de Contrato", "Fecha Inicio", "Fecha Vencimiento",
        "Fecha Creacion", "Vigente"
    ]
    ws.append(headers)

    # Filtrar solo empleados con contrato activo
    empleados = Empleado.objects.filter(
        contratos__activo=True
    ).distinct()

    for empleado in empleados:
        contrato_activo = empleado.contrato_activo  # Asumo que esta propiedad retorna el contrato activo

        ws.append([
            empleado.nombres,
            empleado.apellidos,
            empleado.dpi,
            empleado.tipoc,
            empleado.dcargo,
            contrato_activo.renglon if contrato_activo else "N/A",
            str(contrato_activo.sede.nombre) if contrato_activo and contrato_activo.sede else "N/A",
            str(contrato_activo.puesto.nombre) if contrato_activo and contrato_activo.puesto else "N/A",
            contrato_activo.fecha_inicio.strftime("%d/%m/%Y") if contrato_activo else "N/A",
            contrato_activo.fecha_vencimiento.strftime("%d/%m/%Y") if contrato_activo else "N/A",
            empleado.created_at.strftime("%d/%m/%Y"),
            "Sí" if empleado.tiene_contrato_activo else "No"
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=empleados_contratos_vigentes.xlsx'

    wb.save(response)
    return response

def exportar_empleados_excel_029(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados y Contratos"

    headers = [
        "Nombres", "Apellidos", "DPI", "Cargo", "Detalle del Cargo", "Renglón",
        "Sede", "Puesto de Contrato", "Fecha Inicio", "Fecha Vencimiento", "Fecha Creacion", "Vigente"
    ]
    ws.append(headers)

    empleados = Empleado.objects.filter(
        contratos__activo=True,
        contratos__renglon='029'
    ).distinct()

    for empleado in empleados:
        contrato_activo = empleado.contrato_activo
        ws.append([
            empleado.nombres,
            empleado.apellidos,
            empleado.dpi,
            empleado.tipoc,
            empleado.dcargo,
            contrato_activo.renglon if contrato_activo else "N/A",
            str(contrato_activo.sede) if contrato_activo and contrato_activo.sede else "N/A",
            str(contrato_activo.puesto) if contrato_activo and contrato_activo.puesto else "N/A",
            contrato_activo.fecha_inicio.strftime("%d/%m/%Y") if contrato_activo else "N/A",
            contrato_activo.fecha_vencimiento.strftime("%d/%m/%Y") if contrato_activo else "N/A",
            empleado.created_at.strftime("%d/%m/%Y"),
            "Sí" if empleado.tiene_contrato_activo else "No"
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=empleados_contratos_vigentes_029.xlsx'

    wb.save(response)
    return response


def exportar_empleados_excel_021(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados y Contratos"

    headers = [
        "Nombres", "Apellidos", "DPI", "Cargo", "Detalle del Cargo", "Renglón",
        "Sede", "Puesto de Contrato", "Fecha Inicio", "Fecha Vencimiento", "Fecha Creacion", "Vigente"
    ]
    ws.append(headers)

    empleados = Empleado.objects.filter(
        contratos__activo=True,
        contratos__renglon='021'
    ).distinct()

    for empleado in empleados:
        contrato_activo = empleado.contrato_activo
        ws.append([
            empleado.nombres,
            empleado.apellidos,
            empleado.dpi,
            empleado.tipoc,
            empleado.dcargo,
            contrato_activo.renglon if contrato_activo else "N/A",
            str(contrato_activo.sede) if contrato_activo and contrato_activo.sede else "N/A",
            str(contrato_activo.puesto) if contrato_activo and contrato_activo.puesto else "N/A",
            contrato_activo.fecha_inicio.strftime("%d/%m/%Y") if contrato_activo else "N/A",
            contrato_activo.fecha_vencimiento.strftime("%d/%m/%Y") if contrato_activo else "N/A",
            empleado.created_at.strftime("%d/%m/%Y"),
            "Sí" if empleado.tiene_contrato_activo else "No"
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=empleados_contratos_vigentes_021.xlsx'

    wb.save(response)
    return response

def exportar_empleados_no_vigentes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados y Contratos"

    headers = [
        "Nombres", "Apellidos", "DPI", "Cargo", "Detalle del Cargo", "Renglón",
        "Sede", "Puesto de Contrato", "Fecha Inicio", "Fecha Vencimiento", "Fecha Creacion", "Vigente"
    ]
    ws.append(headers)

    empleados = Empleado.objects.exclude(
        contratos__activo=True
    ).distinct()

    for empleado in empleados:
        contrato_no_vigente = empleado.contratos.filter(activo=False).last()
        ws.append([
            empleado.nombres,
            empleado.apellidos,
            empleado.dpi,
            empleado.tipoc,
            empleado.dcargo,
            contrato_no_vigente.renglon if contrato_no_vigente else "N/A",
            str(contrato_no_vigente.sede) if contrato_no_vigente and contrato_no_vigente.sede else "N/A",
            str(contrato_no_vigente.puesto) if contrato_no_vigente and contrato_no_vigente.puesto else "N/A",
            contrato_no_vigente.fecha_inicio.strftime("%d/%m/%Y") if contrato_no_vigente else "N/A",
            contrato_no_vigente.fecha_vencimiento.strftime("%d/%m/%Y") if contrato_no_vigente else "N/A",
            empleado.created_at.strftime("%d/%m/%Y"),
            "Sí" if empleado.tiene_contrato_activo else "No"
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=empleados_contratos_no_vigentes.xlsx'

    wb.save(response)
    return response




def empleado_detalle(request, id):
    empleado = Empleado.objects.get(id=id)
    
    # Generar el código QR
    qr_url = request.build_absolute_uri()  # Genera la URL completa de la página
    qr_image = qrcode.make(qr_url)

    # Guardar el código QR en un objeto BytesIO
    img_io = BytesIO()
    qr_image.save(img_io, 'PNG')
    img_io.seek(0)

    # Crear un archivo temporal en la memoria
    empleado.qr_code.save(f'qr_{empleado.id}.png', ContentFile(img_io.read()), save=False)
    configuracion = ConfiguracionGeneral.objects.first()
    return render(request, 'empleados/empleado_detalle.html', {
        'empleado': empleado,
        'configuracion': configuracion,
    })