import csv
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.exceptions import ValidationError
from django.db import transaction

from openpyxl import load_workbook

from scompras_app.models import PresupuestoRenglon, Producto, Subproducto


HEADERS_REQUERIDOS = {
    'producto_codigo',
    'subproducto_codigo',
    'codigo_renglon',
    'descripcion',
    'monto_inicial',
}


def to_str(valor):
    if valor is None:
        return ''
    return str(valor).strip()


def zfill_code(valor, width):
    if not valor:
        return ''
    if valor.isdigit():
        return valor.zfill(width)
    return valor


def parse_decimal(valor):
    if valor is None:
        raise ValidationError('El monto inicial es obligatorio.')
    valor_str = to_str(valor)
    if not valor_str:
        raise ValidationError('El monto inicial es obligatorio.')
    valor_str = valor_str.replace(' ', '')
    valor_str = valor_str.replace(',', '')
    try:
        monto = Decimal(valor_str)
    except (InvalidOperation, ValueError):
        raise ValidationError(f'Formato de monto inválido: {valor}')
    return monto


def _normalizar_headers(headers):
    return [str(h).strip().lower() if h is not None else '' for h in headers]


def read_rows(archivo):
    nombre = Path(archivo.name).suffix.lower()
    if nombre == '.csv':
        return _read_csv(archivo)
    if nombre in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        return _read_xlsx(archivo)
    raise ValidationError('Formato de archivo no soportado. Use CSV o XLSX.')


def _read_csv(archivo):
    archivo.seek(0)
    decoded = archivo.read().decode('utf-8-sig').splitlines()
    reader = csv.DictReader(decoded)
    headers = _normalizar_headers(reader.fieldnames or [])
    if not HEADERS_REQUERIDOS.issubset(set(headers)):
        raise ValidationError('El CSV debe contener los encabezados requeridos.')
    mapping = {header: header for header in headers}
    rows = []
    for index, row in enumerate(reader, start=2):
        normalized = {key: row.get(mapping.get(key)) for key in HEADERS_REQUERIDOS}
        rows.append({'numero_fila': index, **normalized})
    return rows


def _read_xlsx(archivo):
    archivo.seek(0)
    workbook = load_workbook(filename=archivo, read_only=True, data_only=True)
    hoja = workbook['Renglones'] if 'Renglones' in workbook.sheetnames else workbook.active
    rows_iter = hoja.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise ValidationError('El archivo XLSX está vacío.')
    headers_normalizados = _normalizar_headers(headers)
    if not HEADERS_REQUERIDOS.issubset(set(headers_normalizados)):
        raise ValidationError('El XLSX debe contener los encabezados requeridos.')
    idx = {header: headers_normalizados.index(header) for header in HEADERS_REQUERIDOS}
    rows = []
    for row_index, row in enumerate(rows_iter, start=2):
        normalized = {key: row[idx[key]] if idx[key] < len(row) else None for key in HEADERS_REQUERIDOS}
        rows.append({'numero_fila': row_index, **normalized})
    return rows


def import_rows(presupuesto, rows, filename, modo='solo_crear'):
    productos = {p.codigo: p for p in Producto.objects.all()}
    subproductos = {
        (s.producto.codigo, s.codigo): s for s in Subproducto.objects.select_related('producto')
    }
    existentes = {
        (r.producto_id, r.subproducto_id, r.codigo_renglon): r
        for r in PresupuestoRenglon.objects.filter(presupuesto_anual=presupuesto).only(
            'id',
            'producto_id',
            'subproducto_id',
            'codigo_renglon',
            'monto_inicial',
            'monto_reservado',
            'monto_ejecutado',
        )
    }

    resultado = {
        'total': 0,
        'creados': 0,
        'duplicados': 0,
        'actualizados': 0,
        'errores': [],
    }

    with transaction.atomic():
        for row in rows:
            resultado['total'] += 1
            fila = row['numero_fila']
            try:
                producto_codigo = to_str(row.get('producto_codigo'))
                subproducto_codigo = zfill_code(to_str(row.get('subproducto_codigo')), 3)
                codigo_renglon = to_str(row.get('codigo_renglon'))
                descripcion = to_str(row.get('descripcion'))
                monto_inicial = parse_decimal(row.get('monto_inicial'))

                if not codigo_renglon:
                    raise ValidationError('El código de renglón es obligatorio.')
                if monto_inicial <= 0:
                    raise ValidationError('El monto inicial debe ser mayor que cero.')

                producto = None
                subproducto = None
                if producto_codigo:
                    producto, creado = Producto.objects.get_or_create(
                        codigo=producto_codigo,
                        defaults={
                            'nombre': f'{producto_codigo}',
                            'descripcion': 'Importado desde carga masiva',
                            'activo': True,
                        },
                    )
                    if creado:
                        productos[producto.codigo] = producto
                if subproducto_codigo:
                    if not producto_codigo:
                        raise ValidationError('Debe indicar producto cuando se especifica subproducto.')
                    subproducto, creado = Subproducto.objects.get_or_create(
                        producto=producto,
                        codigo=subproducto_codigo,
                        defaults={
                            'nombre': f' {subproducto_codigo}',
                            'descripcion': 'Importado desde carga masiva',
                            'activo': True,
                        },
                    )
                    if creado:
                        subproductos[(producto.codigo, subproducto.codigo)] = subproducto

                clave = (producto.id if producto else None, subproducto.id if subproducto else None, codigo_renglon)
                existente = existentes.get(clave)
                if existente:
                    if modo == 'actualizar_si_sin_movimientos':
                        if existente.monto_reservado > 0 or existente.monto_ejecutado > 0:
                            resultado['duplicados'] += 1
                        else:
                            existente.monto_inicial = monto_inicial
                            existente.descripcion = descripcion or existente.descripcion
                            existente.full_clean()
                            existente.save(update_fields=['monto_inicial', 'descripcion', 'fecha_actualizacion'])
                            resultado['actualizados'] += 1
                    else:
                        resultado['duplicados'] += 1
                    continue

                renglon = PresupuestoRenglon(
                    presupuesto_anual=presupuesto,
                    producto=producto,
                    subproducto=subproducto,
                    codigo_renglon=codigo_renglon,
                    descripcion=descripcion or None,
                    monto_inicial=monto_inicial,
                )
                renglon._kardex_referencia = f"Carga masiva ({filename}) fila {fila}"
                renglon.full_clean()
                renglon.save()
                existentes[clave] = renglon
                resultado['creados'] += 1
            except ValidationError as exc:
                resultado['errores'].append(
                    {
                        'fila': fila,
                        'error': '; '.join(exc.messages),
                    }
                )

    return resultado
